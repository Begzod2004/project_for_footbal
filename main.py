from operator import le
import sys
import time
from tkinter import N
import traceback
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from openpyxl import Workbook
from models import Team, Players
from windows.CountriesWindow import CountriesWindow
from windows.TeamWindow import TeamWindow

class Window(QMainWindow):

    def __init__(self) -> None: 
        super().__init__()

        self.initUI()
        self.initActions()
        self.initMenu()
        self.initTable()
        self.fillTable()
        self.msg = QMessageBox()
        # # changing the background color to cyan
        # self.setStyleSheet("background-color: cyan;")
  
        # # creating a label widget
        # self.label = QLabel("Cyan", self)
  
        # # moving position
        # self.label.move(100, 100)
  
        # # setting up border
        # self.label.setStyleSheet("border: 1px solid black;")


      
  
      

    def initUI(self):
        btn_add = QPushButton("Add", self)
        btn_add.move(400, 220)
        btn_add.clicked.connect(self.onAdd)

        btn_add = QPushButton("Update", self)
        btn_add.move(500, 220)
        btn_add.clicked.connect(self.onUpdate)

        btn_add = QPushButton("Delete", self)
        btn_add.move(600, 220)
        btn_add.clicked.connect(self.onDel)

        btn_add = QPushButton("Report",self)
        btn_add.move(700, 220)
        btn_add.clicked.connect(self.onRep)

        ql = QLabel("Name: ", self)
        ql.move(1430, 250)
        ql = QLabel("Age: ", self)
        ql.move(1430, 280)
        ql = QLabel("Goal: ", self)
        ql.move(1430, 305)
        ql = QLabel("Speed: ", self)
        ql.move(1430, 330)

       
        ql = QLabel("Team",self)
        ql.move(1430, 355)
        self.qle_nam = QLineEdit(self)
        self.qle_nam.move(1535, 250)
        self.qle_nam.setMinimumWidth(300)
        self.qle_age = QLineEdit(self)
        self.qle_age.move(1535, 275)
        self.qle_age.setMinimumWidth(300)
        self.qle_gol = QLineEdit(self)
        self.qle_gol.move(1535, 300)
        self.qle_gol.setMinimumWidth(300)
        self.qle_sped = QLineEdit(self)
        self.qle_sped.move(1535, 330)
        self.qle_sped.setMinimumWidth(300)
        self.cbb_team = QComboBox(self)
        self.cbb_team.move(1535, 355)  
        self.cbb_team.setMinimumWidth(300) 


        for item in Team.objects():
            
            self.cbb_team.addItem(item.name,item.id)

    def onRep(self):
        wb = Workbook()
        try:
            ws = wb.active
            ws[f'A{1}'] = "Name"
            ws[f'B{1}'] = "Age"
            ws[f'C{1}'] = "Goal"
            ws[f'D{1}'] = "Speed"
            ws[f'E{1}'] = "Team"
            ws[f'F{1}'] = "Players"

            for sel_row in range(self.table.rowCount()):
                Name = self.table.item(sel_row, 1).text()
                Age = int(self.table.item(sel_row, 2).text())
                Goals = int(self.table.item(sel_row, 3).text())
                Speed = int(self.table.item(sel_row, 4).text())
                tea_name = self.table.item(sel_row, 5).text()
                play_name = self.table.item(sel_row, 7).text()



                ws[f'A{sel_row + 2}'] = Name 
                ws[f'B{sel_row + 2}'] = Age
                ws[f'C{sel_row + 2}'] = Goals
                ws[f'D{sel_row + 2}'] = Speed
                ws[f'E{sel_row + 2}'] = tea_name
                ws[f'F{sel_row + 2}'] = play_name
            # Save the file
            wb.save("sample.xlsx")
            wb.close()

        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()
            wb.close()
            traceback.print_exc()
                

    def onAdd(self):
        try:
            Name = self.qle_nam.text()
            Age = int(self.qle_age.text())
            Goals = int(self.qle_gol.text())
            Speed = int(self.qle_sped.text())
            Team_Id = self.cbb_team.currentData()

            players = Players(Name,Age,Goals,Speed,Team_Id)
            players.save()

            dist = players.Team
            reg = dist.Countries
            row_count = self.table.rowCount()
            self.table.setRowCount(row_count + 1)
            self.table.setItem(row_count, 0,
                               QTableWidgetItem(str(players.id)))
            self.table.setItem(row_count, 1,
                               QTableWidgetItem(players.Name))
            self.table.setItem(row_count, 2,
                               QTableWidgetItem(str(players.Age)))
            self.table.setItem(row_count, 3,
                               QTableWidgetItem(str(players.Goals)))
            self.table.setItem(row_count, 4,
                               QTableWidgetItem(str(players.Speed)))
            self.table.setItem(row_count, 5,
                              
                               QTableWidgetItem(reg.name))
            self.table.setItem(row_count, 6,
                               QTableWidgetItem(str(dist.id)))
            self.table.setItem(row_count, 7,
                               QTableWidgetItem(dist.name))
            self.msg.setIcon(QMessageBox.Information)
            self.msg.setWindowTitle("Bajarildi")
            self.msg.setText("Talaba saqlandi....")
            self.msg.show()
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()

    def onUpdate(self):
        try:
            id = int(self.table.item(self.sel_row, 0).text())
            Name = self.qle_nam.text()
            Age = int(self.qle_age.text())
            Goals = int(self.qle_gol.text())
            Speed = int(self.qle_sped.text())
            Team_Id = self.cbb_team.currentData()

            players = Players(Name, Age, Goals, Speed, Team_Id, id)
            players.save()
            
            dist = players.Team
            reg = dist.Countries
            row_count = self.table.currentRow()
            self.table.setItem(row_count, 0,
                               QTableWidgetItem(str(players.id)))
            self.table.setItem(row_count, 1,
                               QTableWidgetItem(players.Name))
            self.table.setItem(row_count, 2,
                               QTableWidgetItem(str(players.Age)))
            self.table.setItem(row_count, 3,
                               QTableWidgetItem(str(players.Goals)))
            self.table.setItem(row_count, 4,
                               QTableWidgetItem(str(players.Speed)))
            self.table.setItem(row_count, 5,
                              
                               QTableWidgetItem(reg.name))
            self.table.setItem(row_count, 6,
                               QTableWidgetItem(str(dist.id)))
            self.table.setItem(row_count, 7,
                               QTableWidgetItem(dist.name))
            self.msg.setIcon(QMessageBox.Information)
            self.msg.setWindowTitle("Bajarildi")
            self.msg.setText("Talaba saqlandi....")
            self.msg.show()
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()
            traceback.print_exc()
    
    def onDel(self):
        try:

            id = int(self.table.item(self.sel_row, 0).text())
            Name = self.qle_nam.text()
            Age = int(self.qle_age.text())
            Goals = int(self.qle_gol.text())
            Speed = int(self.qle_sped.text())
            Team_Id = int(self.table.item(self.sel_row, 6).text())
            players = Players(Name, Age, Goals, Speed, Team_Id, id)
            players.delete()
            id = self.table.item(self.sel_row, 0).text()
            self.qle_nam.setText('')
            self.qle_age.setText('')
            self.qle_gol.setText('')
            self.qle_sped.setText('')
            

            row_count = self.table.currentRow()
            self.table.removeRow(row_count)

            self.msg.setIcon(QMessageBox.Information)
            self.msg.setWindowTitle("Bajarildi")
            self.msg.setText("Talaba o'chirildi....")
            self.msg.show()
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()

    def initTable(self):
        self.table = QTableWidget(self)  # Создаём таблицу
        self.table.move(400, 250)
        self.table.setMinimumSize(1000,500 )
        self.table.setColumnCount(8)     # Устанавливаем три колонки

        # Устанавливаем заголовки таблицы
        self.table.setHorizontalHeaderLabels(
            ['Id', "   Name   ", "   Age   ", "   Goals   ", "   Speed  ", "   Countries name  ", "  Team id  ", "  Team  "])

        self.table.hideColumn(0)
        self.table.hideColumn(6)
        self.table.clicked.connect(self.onClicked)

    def initActions(self):
        self.newAction = QAction("&New...", self)
        self.newAction.triggered.connect(self.onnewAction)
        self.openAction = QAction("&Open...", self)
        self.saveAction = QAction("&Save", self)
        self.exitAction = QAction("&Exit", self)

        self.regionAction = QAction("&Countries", self)
        self.regionAction.triggered.connect(self.onCountriesWindow)
        self.districtAction = QAction("&Team", self)
        self.districtAction.triggered.connect(self.onTeamWindow)
        
    def fillTable(self):
        for players in Players.objects():
            dist = players.Team
            reg =  dist.Countries
            row_count = self.table.rowCount()
            self.table.setRowCount(row_count + 1)
            self.table.setItem(row_count, 0,
                               QTableWidgetItem(str(players.id)))
            self.table.setItem(row_count, 1,
                               QTableWidgetItem(players.Name))
            self.table.setItem(row_count, 2,
                               QTableWidgetItem(str(players.Age)))
            self.table.setItem(row_count, 3,
                               QTableWidgetItem(str(players.Goals)))
            self.table.setItem(row_count, 4,
                               QTableWidgetItem(str(players.Speed)))
            self.table.setItem(row_count, 5,
                              
                               QTableWidgetItem(str(reg.name)))
            self.table.setItem(row_count, 6,
                               QTableWidgetItem(str(dist.id)))
            self.table.setItem(row_count, 7,
                               QTableWidgetItem(str(dist.name)))
            # делаем ресайз колонок по содержимому
        self.table.resizeColumnsToContents()

    def onClicked(self):
        try:
            self.sel_row = self.table.currentRow()
            Name = self.table.item(self.sel_row, 1).text()    
            self.qle_nam.setText(Name)
            Age = int(self.table.item(self.sel_row, 2).text())
            self.qle_age.setText(str(Age))
            Goals = int(self.table.item(self.sel_row, 3).text())
            self.qle_gol.setText(str(Goals))
            Speed = int(self.table.item(self.sel_row, 4).text())
            self.qle_sped.setText(str(Speed))
            
            Team_Id = self.table.item(self.sel_row, 5).text()
            for i in range(self.cbb_team.count()):
                if self.cbb_team.itemData(i) == Team_Id:
                    self.cbb_team.setCurrentIndex(i)
                    break
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()


    def onnewAction(self):
        pass

    def initMenu(self):
        menuBar = self.menuBar()
        self.setMenuBar(menuBar)

        fileMenu = menuBar.addMenu("&File")
        fileMenu.addAction(self.newAction)
        fileMenu.addAction(self.openAction)
        fileMenu.addAction(self.saveAction)
        fileMenu.addAction(self.exitAction)

        servicesMenu = menuBar.addMenu("&Services")
        servicesMenu.addAction(self.regionAction)
        servicesMenu.addAction(self.districtAction)

        helpMenu = menuBar.addMenu("&Help")

    def onCountriesWindow(self):
        self.regw = CountriesWindow()
        self.regw.show()

    def onTeamWindow(self):
        self.distw = TeamWindow()
        self.distw.show()


app = QApplication(sys.argv)

w = Window()
w.showMaximized()

app.exec()
